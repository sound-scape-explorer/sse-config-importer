from typing import Optional

from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter

from importer.types.ReaderColumnInterface import ReaderColumnInterface
from importer.types.ReaderColumnsInterface import ReaderColumnsInterface


class Reader:
    def __init__(
        self,
        path: str,
        sheet_index: Optional[int] = 0,
    ):
        self.__path = path
        self.__sheet_index = sheet_index

        self.__open()
        self.__read_columns()

    def __open(self):
        self.__workbook = load_workbook(self.__path)
        self.__sheet = self.__workbook.sheetnames[self.__sheet_index]
        self.__worksheet = self.__workbook[self.__sheet]
        self.__max_row = self.__worksheet.max_row

    def __close(self):
        self.__workbook.save(self.__path)

    def __read_columns(self):
        self.__columns: ReaderColumnsInterface = {
            cell.value: {
                'letter': get_column_letter(cell.column),
                'number': cell.column - 1
            } for cell in self.__worksheet[1] if cell.value
        }

    def get_columns(self):
        return self.__columns

    def get_column(self, column_name: str):
        return self.__columns[column_name]

    def read_column(self, column: ReaderColumnInterface):
        results = []

        for row in range(2, self.__max_row + 1):
            cell_name = "{}{}".format(column['letter'], row)
            value = self.__worksheet[cell_name].value
            results.append(value)

        return results
